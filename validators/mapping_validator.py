import os
from pathlib import Path
import openpyxl
from config import settings

class MappingValidator:
    COL_FIELD = "Field"
    COL_DATA_FILE = "Data File"
    COL_COLUMN = "Column"
    COL_NOTES = "Notes/Additional Information"
    COL_CONVERTED = "Converted"

    def __init__(self, mapping_file_path: str = None):
        if mapping_file_path:
            self.file_path = Path(mapping_file_path)
        else:
            # 🎯 Point directly to INGEST directory containing raw data pulled from SharePoint
            ingest_dir = Path(settings.LOCAL_INGEST_DIR)
            
            # Scan for original Excel file (.xlsx) containing 'mapping' in name
            mapping_files = [f for f in ingest_dir.glob("*.xlsx") if "mapping" in f.name.lower()]
            
            if not mapping_files:
                raise FileNotFoundError(
                    f"❌ No original Excel Mapping file (.xlsx) found in ingest directory: {ingest_dir}\n"
                    f"💡 Ensure file name on SharePoint contains the word 'Mapping' (e.g., MediCoop Data Mapping.xlsx)"
                )
            self.file_path = mapping_files[0]

        if not self.file_path.exists():
            raise FileNotFoundError(f"❌ File does not exist: {self.file_path}")

    def validate(self) -> tuple[bool, dict]:
        errors_by_sheet = {}

        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
        except Exception as e:
            return False, {"System": [f"❌ Cannot open Excel Mapping file: {e}"]}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            col_idx_map = {}
            sheet_errors = []
            current_active_data_file = ""
            
            has_header = False
            mapped_field_count = 0

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not any(row): 
                    continue

                row_str = [str(cell).strip() if cell is not None else "" for cell in row]

                # 1. Identify standard Header row
                if self.COL_FIELD in row_str and self.COL_DATA_FILE in row_str:
                    col_idx_map = {
                        self.COL_FIELD: row_str.index(self.COL_FIELD),
                        self.COL_DATA_FILE: row_str.index(self.COL_DATA_FILE),
                        self.COL_COLUMN: row_str.index(self.COL_COLUMN) if self.COL_COLUMN in row_str else -1,
                        self.COL_NOTES: row_str.index(self.COL_NOTES) if self.COL_NOTES in row_str else -1,
                        self.COL_CONVERTED: row_str.index(self.COL_CONVERTED) if self.COL_CONVERTED in row_str else -1,
                    }
                    current_active_data_file = ""
                    has_header = True
                    continue

                # 2. Process Data rows
                if col_idx_map:
                    field_val = row_str[col_idx_map[self.COL_FIELD]]
                    if not field_val or field_val == self.COL_FIELD: 
                        continue

                    data_file_val = row_str[col_idx_map[self.COL_DATA_FILE]]
                    column_val = row_str[col_idx_map[self.COL_COLUMN]] if col_idx_map[self.COL_COLUMN] != -1 else ""
                    notes_val = row_str[col_idx_map[self.COL_NOTES]] if col_idx_map[self.COL_NOTES] != -1 else ""
                    converted_val = row_str[col_idx_map[self.COL_CONVERTED]] if col_idx_map[self.COL_CONVERTED] != -1 else ""

                    if data_file_val:
                        current_active_data_file = data_file_val
                    effective_data_file = data_file_val or current_active_data_file

                    is_converted = converted_val.lower() in ["yes", "y", "true", "1"]
                    has_source_mapping = bool(effective_data_file and column_val)
                    has_notes = bool(notes_val)
                    
                    if is_converted or has_source_mapping or has_notes:
                        mapped_field_count += 1

                    # Check Error 1: Column declared but missing Data File
                    if column_val and not effective_data_file:
                        sheet_errors.append(
                            f"Row {row_idx:04d} | Field '{field_val}': Has Column='{column_val}' but Data File not declared on this or above rows!"
                        )

                    # Check Error 2: Converted = Yes but completely empty
                    if is_converted and not has_source_mapping and not has_notes:
                        sheet_errors.append(
                            f"Row {row_idx:04d} | Field '{field_val}': Marked Converted='Yes' but MISSING Mapping (Data File/Column) and Notes!"
                        )

            # Check Error 3: Sheet completely empty (Requires N/A in Notes column to skip)
            if has_header and mapped_field_count == 0:
                sheet_errors.append(
                    f"Entire Sheet is left blank (No Fields mapped). "
                    f"💡 Solution: If Credit Union DOES NOT USE this module, type 'N/A' in Notes column of any row to confirm!"
                )

            if sheet_errors:
                errors_by_sheet[sheet_name] = sheet_errors

        wb.close()
        return len(errors_by_sheet) == 0, errors_by_sheet