import sqlite3
import sys
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "workspace" / "rules.db"
REPORT_DIR = BASE_DIR / "workspace" / "reports"


def get_available_cu_ids(db_path: Path) -> list[str]:
    """Retrieve distinct non-global CU IDs stored in the rule_store database."""
    if not db_path.exists():
        return []
    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT DISTINCT cu_id FROM rule_store WHERE cu_id IS NOT NULL AND cu_id != '' AND is_global != 1"
        )
        rows = cursor.fetchall()
        return [r[0] for r in rows if r[0]]


def export_rule_verification_report(cu_id: str) -> Path:
    """Export Rule Verification Report properly sorted according to Mapping file for a specific CU."""
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_file = REPORT_DIR / f"Rule_Verification_Report_{cu_id}.xlsx"

    if not DB_PATH.exists():
        print(f"❌ Database not found at {DB_PATH}. Please run rule_engine.py first!")
        return report_file

    query = """
        SELECT 
            id AS "Rule_ID",
            cu_id AS "CU_ID",
            sheet_name AS "Sheet_Name",
            section_name AS "Section_Name",
            target_field AS "Target_Field",
            data_file AS "Source_File",
            column_letter AS "Source_Col",
            raw_notes AS "Raw_Notes",
            dsl_readable AS "Draft_Rule_DSL",
            rule_type AS "Rule_Type",
            parsed_by AS "Parsed_By",
            status AS "Current_Status"
        FROM rule_store
        WHERE cu_id = ? OR is_global = 1
        ORDER BY id ASC
    """

    with sqlite3.connect(DB_PATH) as conn:
        df = pd.read_sql_query(query, conn, params=(cu_id,))

    if df.empty:
        print(f"⚠️ No Rule data in DB to export report for CU: [{cu_id}]!")
        return report_file

    # Add columns for QA assessment
    df["Decision (QA)"] = ""        # APPROVE / EDIT / REJECT
    df["QA Edited DSL"] = ""        # Input new DSL if EDIT is selected
    df["QA Reviewer"] = ""          # QA Reviewer Name (e.g., QA_John)
    df["QA Notes"] = ""             # QA Notes/Comments

    # Auto-suggest Decision for high-confidence rules
    df.loc[df["Current_Status"] == "AUTO_PARSED", "Decision (QA)"] = "APPROVE"
    df.loc[df["Current_Status"] == "NEEDS_REVIEW", "Decision (QA)"] = ""

    # Re-order columns for best QA visibility
    ordered_cols = [
        "Rule_ID", "Section_Name", "Target_Field", "Source_File", "Source_Col",
        "Raw_Notes", "Draft_Rule_DSL", "Decision (QA)", "QA Edited DSL", 
        "QA Reviewer", "QA Notes", "Rule_Type", "Current_Status"
    ]
    df = df[ordered_cols]

    # Export to Excel and apply visual formatting
    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Rule_Verification", index=False)
        
        ws = writer.sheets["Rule_Verification"]
        
        # Header formatting
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        # Section Rule row formatting
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(name="Calibri", size=10, bold=True, color="002060")

        # Format QA Decision cells with light yellow background
        qa_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        ws.freeze_panes = "A2"  # Freeze header row

        for col_num, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Coloring and aligning each row
        for row_idx, row_data in enumerate(df.itertuples(), start=2):
            target_field = str(row_data.Target_Field)
            is_section_rule = (target_field == "_SECTION_RULE_")

            for col_idx in range(1, len(ordered_cols) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Highlight Decision & QA Edited cells
                col_name = ordered_cols[col_idx - 1]
                if col_name in ["Decision (QA)", "QA Edited DSL"]:
                    cell.fill = qa_fill

                # If Section Rule row (Red cell/Table filter) -> Apply gray border
                if is_section_rule:
                    cell.fill = section_fill
                    cell.font = section_font

        # Auto-adjust column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    print(f"📊 Successfully generated Verification Report for CU [{cu_id}] at:")
    print(f"   👉 {report_file.resolve()}\n")
    return report_file


def apply_qa_decisions(reviewed_report_path: Path):
    """Read QA reviewed Excel file -> Update SQLite & Save Audit History to rule_history table."""
    if not reviewed_report_path.exists():
        print(f"❌ Reviewed report file not found: {reviewed_report_path}")
        return

    df = pd.read_excel(reviewed_report_path, sheet_name="Rule_Verification")
    
    stats = {"approved": 0, "edited": 0, "rejected": 0, "skipped": 0}

    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()

        for idx, row in df.iterrows():
            rule_id = int(row["Rule_ID"])
            decision = str(row.get("Decision (QA)", "")).strip().upper()
            qa_edited_dsl = str(row.get("QA Edited DSL", "")).strip()
            reviewer = str(row.get("QA Reviewer", "")).strip() or "QA_USER"
            qa_notes = str(row.get("QA Notes", "")).strip()

            if pd.isna(decision) or not decision or decision not in ["APPROVE", "EDIT", "REJECT"]:
                stats["skipped"] += 1
                continue

            cursor.execute("SELECT dsl_readable, cu_id, sheet_name, section_name, target_field FROM rule_store WHERE id = ?", (rule_id,))
            curr_rule = cursor.fetchone()
            if not curr_rule:
                continue

            prev_dsl, cu_id, sheet_name, sec_name, target_field = curr_rule

            new_dsl = prev_dsl
            new_status = "APPROVED"

            if decision == "APPROVE":
                new_status = "VERIFIED_APPROVED"
                stats["approved"] += 1

            elif decision == "EDIT":
                new_status = "VERIFIED_EDITED"
                new_dsl = qa_edited_dsl if qa_edited_dsl else prev_dsl
                stats["edited"] += 1

            elif decision == "REJECT":
                new_status = "REJECTED"
                stats["rejected"] += 1

            cursor.execute("""
                UPDATE rule_store 
                SET dsl_readable = ?, status = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (new_dsl, new_status, rule_id))

            cursor.execute("""
                INSERT INTO rule_history 
                (rule_id, cu_id, sheet_name, section_name, target_field, action, previous_dsl, new_dsl, reviewer, review_notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (rule_id, cu_id, sheet_name, sec_name, target_field, decision, prev_dsl, new_dsl, reviewer, qa_notes))

        conn.commit()

    print("=" * 50)
    print("✅ APPLIED QA DECISIONS TO DATABASE:")
    print(f"   - Approved: {stats['approved']} rules")
    print(f"   - Edited: {stats['edited']} rules")
    print(f"   - Rejected: {stats['rejected']} rules")
    print(f"   - Skipped (No Decision input): {stats['skipped']} rules")
    print("=" * 50 + "\n")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        arg_val = sys.argv[1]
        
        # If user passes an existing excel file path or 'apply' command
        if Path(arg_val).exists() and arg_val.endswith(".xlsx"):
            apply_qa_decisions(Path(arg_val))
        elif arg_val.lower() == "apply" and len(sys.argv) > 2:
            apply_qa_decisions(Path(sys.argv[2]))
        else:
            # Treat argument as CU_ID to export report
            export_rule_verification_report(cu_id=arg_val.upper())
    else:
        # Auto-detect CU IDs from database and export report for each
        available_cus = get_available_cu_ids(DB_PATH)
        if not available_cus:
            print("❌ No rule data found in DB. Please run rule_engine.py first!")
        else:
            for target_cu in available_cus:
                export_rule_verification_report(cu_id=target_cu)